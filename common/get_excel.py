"""
author：lulu
time:2020/12/17  11:23

"""
import openpyxl


class RW_excel():

    def __init__(self, filename, sheetname):
        """
        初始化文件名和表单名
        :param filename: excel文件名（xlsx文件）
        :param sheetname: 表单名
        """
        self.filename = filename
        self.sheetname = sheetname

    def r_excel(self):
        """
        读取excel文件
        :return: 列表内返回的数据格式为标题+数据的字典：
                [{标题1：数据1，标题2：数据2},{标题1：数据3，标题2：数据4}]
        """
        wookbook = openpyxl.load_workbook(self.filename)
        sheet = wookbook[self.sheetname]
        sheet_value = list(sheet.rows)
        title = [i.value for i in sheet_value[0]]
        data_list = []
        for i in sheet_value[1:]:
            data = [j.value for j in i]
            param = dict(zip(title, data))
            data_list.append(param)
        return data_list

    def w_excel(self, row, cell, value):
        """
        往指定的cell内写入数据
        :param row: 写入的行
        :param cell: 写入的列
        :param value: 写入的数据
        :return:
        """
        wookbook = openpyxl.load_workbook(self.filename)
        sheet = wookbook[self.sheetname]
        sheet.cell(row=row, column=cell, value=value)
        wookbook.save(self.filename)

# if __name__ == '__main__':
#     a=RW_excel()
#     a.r_excel()
